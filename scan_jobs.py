from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 Chrome/127 Safari/537.36"
)

JOB_LINK_HINTS = (
    "job", "jobs", "career", "careers", "opening", "position", "role",
    "greenhouse.io", "lever.co", "workdayjobs", "myworkdayjobs",
    "smartrecruiters", "ashbyhq", "icims", "jobvite",
)

BLOCKED_LINK_HINTS = (
    "linkedin.com", "glassdoor.", "indeed.", "naukri.", "monster.",
    "ambitionbox.", "facebook.", "instagram.", "youtube.", "google.com/search",
    "bing.com/search", "mailto:", "tel:",
)


@dataclass
class Company:
    sheet: str
    company_type: str
    name: str
    career_page: str
    status: str


@dataclass
class JobMatch:
    run_date: str
    company_name: str
    company_type: str
    career_page: str
    job_name: str
    job_link: str
    location: str
    matched_title: str
    matched_skills: str
    experience_match: str
    match_score: int
    status: str
    first_seen_date: str
    last_seen_date: str


def norm(value: Any) -> str:
    return str(value or "").strip()


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def lower_list(values: list[str]) -> list[str]:
    return [str(v).strip().lower() for v in values if str(v).strip()]


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_companies(workbook_path: Path, config: dict[str, Any]) -> list[Company]:
    allowed = set(config["scan"]["allowed_statuses"])
    skip_url_contains = lower_list(config["scan"].get("skip_url_contains", []))
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    companies: list[Company] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        company_type = "Product-Based" if "Product" in sheet_name else "Service-Based"
        headers = [norm(cell.value) for cell in ws[4]]
        try:
            company_col = headers.index("Company Name") + 1
            url_col = headers.index("Career Page") + 1
            status_col = headers.index("Status") + 1
        except ValueError:
            continue

        for row in ws.iter_rows(min_row=5, values_only=False):
            name = norm(row[company_col - 1].value)
            career_page = norm(row[url_col - 1].value)
            status = norm(row[status_col - 1].value)
            if not name or not career_page or status not in allowed:
                continue
            if any(skip in career_page.lower() for skip in skip_url_contains):
                continue
            companies.append(Company(sheet_name, company_type, name, career_page, status))

    return companies


def fetch_html(url: str, timeout: int) -> tuple[str, str]:
    response = requests.get(
        url,
        timeout=timeout,
        allow_redirects=True,
        headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml"},
    )
    response.raise_for_status()
    content_type = response.headers.get("content-type", "")
    if "text/html" not in content_type and "application/xhtml" not in content_type:
        return response.url, ""
    return response.url, response.text


def is_probable_job_link(url: str, text: str) -> bool:
    value = f"{url} {text}".lower()
    if any(blocked in value for blocked in BLOCKED_LINK_HINTS):
        return False
    return any(hint in value for hint in JOB_LINK_HINTS)


def extract_job_candidates(company: Company, html: str, final_url: str, max_jobs: int) -> list[dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    candidates: dict[str, dict[str, str]] = {}

    for script in soup(["script", "style", "noscript"]):
        script.decompose()

    for a in soup.find_all("a", href=True):
        text = clean_text(a.get_text(" "))
        href = urljoin(final_url, a["href"])
        if not text or len(text) > 140:
            continue
        if not is_probable_job_link(href, text):
            continue

        container = a
        for _ in range(3):
            if container.parent:
                container = container.parent
        context = clean_text(container.get_text(" "))[:1200]
        link_key = normalize_job_link(href)
        candidates[link_key] = {
            "title": text,
            "link": href,
            "context": context,
            "location": guess_location(context),
        }
        if len(candidates) >= max_jobs:
            break

    return list(candidates.values())


def normalize_job_link(url: str) -> str:
    parsed = urlparse(url)
    return parsed._replace(fragment="", query=parsed.query[:300]).geturl().rstrip("/")


def guess_location(text: str) -> str:
    known = [
        "Bangalore", "Bengaluru", "India", "Remote", "Hybrid", "Hyderabad",
        "Chennai", "Pune", "Mumbai", "Noida", "Gurugram", "Gurgaon",
    ]
    found = [loc for loc in known if re.search(rf"\b{re.escape(loc)}\b", text, re.I)]
    return ", ".join(dict.fromkeys(found))


def match_job(job: dict[str, str], config: dict[str, Any]) -> dict[str, Any] | None:
    title = job["title"]
    text = f"{job['title']} {job.get('context', '')}".lower()
    title_lower = title.lower()

    primary_titles = lower_list(config["role_focus"]["primary"])
    secondary_titles = lower_list(config["role_focus"]["secondary"])
    required = lower_list(config["skills"]["required_preferred"])
    strong = lower_list(config["skills"]["strong_boost"])
    supporting = lower_list(config["skills"]["supporting"])
    locations = lower_list(config["locations"])
    exp_include = lower_list(config["experience"]["include"])
    exp_exclude = lower_list(config["experience"]["exclude"])

    excluded = [word for word in exp_exclude if word in text]
    if excluded:
        return None

    matched_title = ""
    score = 0
    for keyword in primary_titles:
        if keyword in title_lower:
            matched_title = keyword
            score += 55
            break
    if not matched_title:
        for keyword in secondary_titles:
            if keyword in title_lower:
                matched_title = keyword
                score += 30
                break

    matched_required = [skill for skill in required if re.search(rf"\b{re.escape(skill)}\b", text, re.I)]
    matched_strong = [skill for skill in strong if skill in text]
    matched_supporting = [skill for skill in supporting if skill in text]
    matched_exp = [exp for exp in exp_include if exp in text]
    matched_locations = [loc for loc in locations if loc in text]

    if matched_required:
        score += 35
    elif "java" in title_lower:
        matched_required = ["java"]
        score += 35
    elif matched_title in {"backend developer", "software engineer", "associate software engineer"}:
        score -= 20

    score += min(35, 8 * len(matched_strong))
    score += min(12, 3 * len(matched_supporting))
    score += 12 if matched_exp else 0
    score += 10 if matched_locations else 0

    if score < 55:
        return None

    all_skills = matched_required + matched_strong + matched_supporting
    return {
        "matched_title": matched_title or "java/backend keyword",
        "matched_skills": ", ".join(dict.fromkeys(all_skills)),
        "experience_match": ", ".join(dict.fromkeys(matched_exp)) or "Not specified",
        "location": job.get("location") or ", ".join(dict.fromkeys(matched_locations)) or "Not specified",
        "score": score,
    }


def job_id(company_name: str, job_name: str, job_link: str) -> str:
    raw = f"{company_name}|{job_name}|{normalize_job_link(job_link)}".lower()
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def load_history(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [norm(v) for v in rows[0]]
    history = {}
    for row in rows[1:]:
        record = {headers[i]: norm(row[i]) if i < len(row) else "" for i in range(len(headers))}
        key = record.get("Job ID")
        if key:
            history[key] = record
    return history


def write_workbook(path: Path, rows: list[dict[str, Any]], sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = [
        "Job ID", "Run Date", "Company Name", "Company Type", "Career Page",
        "Job Name", "Job Link", "Location", "Matched Skills", "Matched Title",
        "Experience Match", "Match Score", "Status", "First Seen Date", "Last Seen Date",
    ]
    ws.append(headers)
    for item in rows:
        ws.append([item.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    widths = {
        "A": 18, "B": 14, "C": 28, "D": 18, "E": 45, "F": 42, "G": 55,
        "H": 24, "I": 35, "J": 28, "K": 24, "L": 12, "M": 16, "N": 16, "O": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def chunks(items: list[Company], size: int) -> list[list[Company]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


def scan_company(
    company: Company,
    config: dict[str, Any],
    timeout: int,
    max_jobs: int,
    today: str,
    history: dict[str, dict[str, str]],
) -> tuple[list[dict[str, Any]], set[str], dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    seen_today: set[str] = set()

    try:
        final_url, html = fetch_html(company.career_page, timeout)
        candidates = extract_job_candidates(company, html, final_url, max_jobs)
        for candidate in candidates:
            match = match_job(candidate, config)
            if not match:
                continue
            key = job_id(company.name, candidate["title"], candidate["link"])
            previous = history.get(key)
            seen_today.add(key)
            status = "Still Open" if previous else "New"
            first_seen = previous.get("First Seen Date", today) if previous else today
            matches.append({
                "Job ID": key,
                "Run Date": today,
                "Company Name": company.name,
                "Company Type": company.company_type,
                "Career Page": company.career_page,
                "Job Name": candidate["title"],
                "Job Link": candidate["link"],
                "Location": match["location"],
                "Matched Skills": match["matched_skills"],
                "Matched Title": match["matched_title"],
                "Experience Match": match["experience_match"],
                "Match Score": match["score"],
                "Status": status,
                "First Seen Date": first_seen,
                "Last Seen Date": today,
            })
        log = {"company": company.name, "status": "ok", "jobs_seen": len(candidates), "matches": len(matches)}
    except Exception as exc:
        log = {"company": company.name, "status": "error", "error": str(exc), "url": company.career_page}

    return matches, seen_today, log


def scan(config_path: Path, dry_run: bool = False) -> int:
    config = load_config(config_path)
    workbook_path = BASE_DIR / config["input_workbook"]
    outputs_dir = BASE_DIR / config["outputs_dir"]
    today = date.today().isoformat()
    history_path = outputs_dir / config["history_workbook"]
    report_path = outputs_dir / f"job_matches_{today}.xlsx"
    log_path = outputs_dir / f"scan_log_{today}.json"

    companies = read_companies(workbook_path, config)
    max_companies = int(config["scan"].get("max_companies", 0) or 0)
    if max_companies > 0:
        companies = companies[:max_companies]

    timeout = int(config["scan"]["request_timeout_seconds"])
    delay = float(config["scan"]["delay_between_companies_seconds"])
    max_jobs = int(config["scan"]["max_jobs_per_company"])
    worker_count = max(1, int(config["scan"].get("worker_count", 1) or 1))
    batch_size = max(1, int(config["scan"].get("batch_size", worker_count) or worker_count))
    history = load_history(history_path)
    seen_today: set[str] = set()
    matches: list[dict[str, Any]] = []
    logs: list[dict[str, Any]] = []

    company_batches = chunks(companies, batch_size)
    processed = 0
    print(f"Scanning {len(companies)} companies in {len(company_batches)} batches with {worker_count} workers.")

    for batch_index, batch in enumerate(company_batches, start=1):
        print(f"Batch {batch_index}/{len(company_batches)}: {len(batch)} companies")
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = {
                executor.submit(scan_company, company, config, timeout, max_jobs, today, history): company
                for company in batch
            }
            for future in as_completed(futures):
                company = futures[future]
                company_matches, company_seen, log = future.result()
                matches.extend(company_matches)
                seen_today.update(company_seen)
                logs.append(log)
                processed += 1
                print(f"[{processed}/{len(companies)}] {company.name}: {log['status']} ({len(company_matches)} matches)")
        if delay > 0 and batch_index < len(company_batches):
            time.sleep(delay)

    combined = {key: value for key, value in history.items()}
    for row in matches:
        combined[row["Job ID"]] = row
    for key, row in list(combined.items()):
        if key not in seen_today and row.get("Last Seen Date") != today:
            row["Status"] = "Not Seen Today"

    sorted_matches = sorted(matches, key=lambda r: (-int(r["Match Score"]), r["Company Name"], r["Job Name"]))
    sorted_history = sorted(combined.values(), key=lambda r: (r.get("Company Name", ""), r.get("Job Name", "")))

    if dry_run:
        print(json.dumps({"companies": len(companies), "matches": len(sorted_matches), "history_rows": len(sorted_history)}, indent=2))
        return 0

    write_workbook(report_path, sorted_matches, "Daily Matches")
    write_workbook(history_path, sorted_history, "Job History")
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text(json.dumps(logs, indent=2), encoding="utf-8")
    print(f"Saved daily report: {report_path}")
    print(f"Saved history: {history_path}")
    print(f"Saved scan log: {log_path}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan company career pages for matching Java/backend jobs.")
    parser.add_argument("--config", default="job_search_config.json", help="Path to JSON config file.")
    parser.add_argument("--dry-run", action="store_true", help="Scan and print counts without writing Excel outputs.")
    args = parser.parse_args()
    return scan(BASE_DIR / args.config, dry_run=args.dry_run)


if __name__ == "__main__":
    sys.exit(main())
